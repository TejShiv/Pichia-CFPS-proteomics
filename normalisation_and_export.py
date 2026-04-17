# normalisation_and_export.py
#
# Shivakumar et al. — Opening the black box: proteomic and single-vesicle dissection of microsomes in a Pichia pastoris cell-free system
#
# Applies intensity normalisation to LC-MS/MS data from
# K. phaffii SuperMan5 cell-free lysate fractions and exports the
# normalised protein table together with all panel-level data to
# a structured Excel workbook.
#
# Input:  Proteomics_raw_data.xlsx  (687 protein entries x 6 fractions)
# Output: Main_Figure_Panel_Data.xlsx (7-sheet workbook)
#
# Normalisation steps
# -------------------
# Step 1 — Loading correction
#   Raw intensities are divided by the total protein loaded per fraction
#   (determined by BCA assay) and rescaled to a 50 µg reference. This
#   corrects for unequal input across fractions.
#
# Step 2 — Median normalisation
#   Each fraction is scaled so its median intensity across all detected
#   proteins equals the global median across all six fractions. This
#   corrects for residual run-to-run LC-MS variation.
#
# Functional annotation
# ---------------------
# Proteins are assigned to GO Slim categories by a three-tiered strategy:
#   (i)  Direct lookup against a curated K. phaffii gene-symbol map
#   (ii) Regex matching against UniProt protein descriptions
#   (iii) Default assignment to "Other" for unannotated entries
#
# Dependencies: pandas, numpy, openpyxl
# Usage:        python normalisation_and_export.py

import re
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

DATA_PATH = "Proteomics_raw_data.xlsx"
OUT_PATH  = "Main_Figure_Panel_Data.xlsx"

# Total protein loaded per fraction, determined by BCA assay (µg)
PROTEIN_LOADED_UG = {
    "BfA_lysate":       46,
    "BfA_supernatant":  23,
    "BfA_microsomes":  250,
    "lysate":           44,
    "supernatant":      31,
    "microsomes":       98,
}
REFERENCE_UG = 50  # intensities are scaled as if this amount were loaded

QUANT_COLS = ["BfA_lysate", "BfA_supernatant", "BfA_microsomes",
              "lysate", "supernatant", "microsomes"]
CTRL_COLS  = ["lysate", "supernatant", "microsomes"]


# ---------------------------------------------------------------------------
# Functional annotation
# Curated GO Slim map: K. phaffii gene symbol -> functional category.
# Entries cover all proteins detected across all six fractions in this study.
# For proteins absent from this map, a regex fallback is applied to the
# UniProt protein description; remaining unannotated proteins are labelled
# "Other".
# ---------------------------------------------------------------------------

GO_SLIM_MAP = {
    **{g: "Translation/Ribosomal" for g in [
        "RPS17B","RPS12","RPS13","RPS7","RPS7B","RPS19B","RPS1","RPS4A",
        "RPS4B","RPS8A","RPS9A","RPS9B","EFT1","EFT2","EFB1","TEF1","TEF2",
        "TEF4","YEF3","SUI1","TIF1","TIF2","STM1","ASC1","EGD1","EGD2",
        "RPL5","RPL7A","RPL10","RPP2B","RPP0",
    ]},
    **{g: "Energy metabolism" for g in [
        "TDH1","TDH2","TDH3","ENO1","ENO2","PGK1","GPM1","FBA1","TPI1",
        "PGI1","GLK1","HXK1","HXK2","CDC19","PDC1","PDC5","PDC6",
        "ADH1","ADH2","ADH3","GPD1","GPD2","GDH1","GDH2","GDH3","LPD1",
        "ACO1","CIT1","MDH1","MDH2","COR1","TKL1","TAL1","ZWF1","GND1",
        "GND2","INO1","ERG10","ATP1","ATP2",
    ]},
    **{g: "Chaperones/Protein folding" for g in [
        "KAR2","PDI1","ERO1","CPR1","CPR2","CPR3","SSA1","SSA2","SSA3",
        "SSA4","SSB1","SSB2","SSC1","SSZ1","SSE1","SSE2","STI1","HSP82",
        "HSC82","TCP1","CCT2","CCT3","CCT4","CCT5","CCT6","CCT7","CCT8",
        "HSP60","HSP10","HSP104","HSP26","HSP42","HSP12",
    ]},
    **{g: "ER/Golgi/Secretory" for g in [
        "SEC61","SEC62","SEC63","SEC11","SEC13","SEC16","SEC17","SEC18",
        "SEC23","SEC24","SEC31","SEC39","DSL1","SRP14","SRP54","SRP68",
        "OST1","STT3","WBP1","DPM1","SEC53","OCH1","MNS1","ALG1","ALG2",
        "CDC48","IRE1","ARL3","PIK1","STE13","LDB19","APM3","APL6",
    ]},
    **{g: "Amino acid metabolism" for g in [
        "MET6","CYS3","ARG2","ARO8","LYS9","ILV5","CAR1","CAR2","SAH1",
        "MET2","LEU1","LEU2","PRO1","PRO2","TRP1","TRP2","TRP3",
    ]},
    **{g: "Redox homeostasis" for g in [
        "TRR1","TRR2","TRX1","TRX2","TSA1","TSA2","AHP1","GRX1","GRX2",
        "GLR1","SOD1","SOD2","CTT1",
    ]},
    **{g: "Proteolysis/Proteasome" for g in [
        "PRE1","PRE2","PRE3","PRE4","PRE5","PRE6","RPN1","RPN2","RPN3",
        "RPT1","RPT2","RPT3","RPT4","RPT5","RPT6","UBI4","DOA4",
    ]},
    **{g: "Nucleic acid metabolism" for g in [
        "ADO1","GUA1","ADE17","YNK1","GND2","URA1","URA2","URA3","RIB1",
    ]},
}


def get_go_slim(gene, desc):
    """Return GO Slim category for a gene symbol, with description fallback."""
    g = str(gene).upper().strip()
    if g in GO_SLIM_MAP:
        return GO_SLIM_MAP[g]
    d = str(desc).lower()
    if re.search(r"ribosom|elongation factor|translation factor", d):
        return "Translation/Ribosomal"
    if re.search(r"glyceraldeh|enolase|phosphoglycerate|pyruvate|aldolase|"
                 r"hexokinase|dehydrogenase|transketol|citrate synth|oxidoreduct", d):
        return "Energy metabolism"
    if re.search(r"chaperone|heat.shock|hsp\d|grp\d|cyclophilin|disulfide.isomerase", d):
        return "Chaperones/Protein folding"
    if re.search(r"endoplasmic|golgi|sec\d|secretory|translocon|n.glycan|"
                 r"mannosyl|glycosyl|dolichol|vesicle|coatomer", d):
        return "ER/Golgi/Secretory"
    if re.search(r"aminotransfer|amino acid biosyn|methionine|arginine|lysine", d):
        return "Amino acid metabolism"
    if re.search(r"peroxiredoxin|thioredoxin reduct|glutaredoxin|superoxide|catalase", d):
        return "Redox homeostasis"
    if re.search(r"proteasome|ubiquitin|protease|peptidase", d):
        return "Proteolysis/Proteasome"
    if re.search(r"nucleotide|purine|pyrimidine|riboflavin|rna polymerase", d):
        return "Nucleic acid metabolism"
    if re.search(r"uncharacterized|hypothetical|unknown", d):
        return "Uncharacterised"
    return "Other"


# ---------------------------------------------------------------------------
# Secretory pathway reference panel
# 39 proteins spanning nine functional groups, used for Table 1 in the
# manuscript. Detection is assessed per fraction after normalisation.
# ---------------------------------------------------------------------------

SECRETORY_PROTEINS = [
    ("SEC61",  "SEC61 (translocon α)",              "ER translocon"),
    ("SEC62",  "SEC62 (translocon)",                 "ER translocon"),
    ("SEC63",  "SEC63 (translocon)",                 "ER translocon"),
    ("SEC11",  "SEC11 (signal peptidase)",           "ER translocon"),
    ("SRP54",  "SRP54",                              "SRP pathway"),
    ("SRP14",  "SRP14",                              "SRP pathway"),
    ("SRP68",  "SRP68",                              "SRP pathway"),
    ("KAR2",   "KAR2 (BiP/Hsp70)",                  "ER chaperone"),
    ("PDI1",   "PDI1 (protein disulfide isomerase)", "ER chaperone"),
    ("ERO1",   "ERO1",                               "ER chaperone"),
    ("CPR1",   "CPR1 (Cyclophilin)",                 "ER chaperone"),
    ("STT3",   "STT3 (OST catalytic subunit)",       "N-glycosylation"),
    ("OST1",   "OST1",                               "N-glycosylation"),
    ("WBP1",   "WBP1",                               "N-glycosylation"),
    ("DPM1",   "DPM1 (Dol-P-Man synthase)",          "N-glycosylation"),
    ("SEC53",  "SEC53 (phosphomannomutase)",          "N-glycosylation"),
    ("OCH1",   "OCH1 (α-1,6-mannosyltransferase)",   "N-glycosylation"),
    ("MNS1",   "MNS1 (α-1,2-mannosidase)",           "N-glycosylation"),
    ("CDC48",  "CDC48 (p97/VCP)",                    "ERAD"),
    ("IRE1",   "IRE1 (UPR sensor kinase)",           "ERAD"),
    ("SEC23",  "SEC23 (COPII inner coat)",            "COPII vesicles"),
    ("SEC24",  "SEC24 (COPII inner coat)",            "COPII vesicles"),
    ("SEC13",  "SEC13 (COPII outer coat)",            "COPII vesicles"),
    ("ARL3",   "ARL3 (Golgi GTPase)",                "Golgi/trafficking"),
    ("PIK1",   "PIK1 (PI4-kinase)",                  "Golgi/trafficking"),
    ("STE13",  "STE13 (DPAP-A)",                     "Golgi/trafficking"),
    ("DSL1",   "DSL1 (ER tethering complex)",        "Golgi/trafficking"),
    ("SEC39",  "SEC39 (DSL complex subunit)",         "Golgi/trafficking"),
    ("EGD1",   "EGD1 (NACβ subunit)",                "Co-translational"),
    ("SSZ1",   "SSZ1 (Hsp70/RAC complex)",           "Cytosolic chaperone"),
    ("SSB2",   "SSB2 (Hsp70/RAC complex)",           "Cytosolic chaperone"),
    ("SSA1",   "SSA1 (Hsp70)",                       "Cytosolic chaperone"),
    ("SSA3",   "SSA3 (Hsp70)",                       "Cytosolic chaperone"),
    ("SSC1",   "SSC1 (mitochondrial Hsp70)",          "Cytosolic chaperone"),
    ("SSE1",   "SSE1 (Hsp110)",                      "Cytosolic chaperone"),
    ("HSP82",  "HSP82 (Hsp90)",                      "Cytosolic chaperone"),
    ("STI1",   "STI1 (Hop co-chaperone)",             "Cytosolic chaperone"),
    ("TCP1",   "TCP1 (TRiC/CCT complex)",             "Cytosolic chaperone"),
]

GROUP_ORDER   = list(dict.fromkeys(g for _, _, g in SECRETORY_PROTEINS))
GROUP_COLOURS = {
    "ER translocon":       "C0392B",
    "SRP pathway":         "E74C3C",
    "ER chaperone":        "E67E22",
    "N-glycosylation":     "F48FB1",
    "ERAD":                "9B59B6",
    "COPII vesicles":      "2980B9",
    "Golgi/trafficking":   "1ABC9C",
    "Co-translational":    "3498DB",
    "Cytosolic chaperone": "27AE60",
}

CAT_COLOURS = {
    "Translation/Ribosomal":     "1B4F72",
    "Energy metabolism":          "F0B27A",
    "Chaperones/Protein folding": "27AE60",
    "Redox homeostasis":          "6C3483",
    "Proteolysis/Proteasome":     "1ABC9C",
    "Amino acid metabolism":      "922B21",
    "ER/Golgi/Secretory":         "E74C3C",
    "Nucleic acid metabolism":    "00BCD4",
    "Uncharacterised":            "A9CCE3",
    "Other":                      "95A5A6",
}


# ---------------------------------------------------------------------------
# Spreadsheet style helpers
# ---------------------------------------------------------------------------

def _fill(hex_col):
    return PatternFill("solid", fgColor=hex_col.lstrip("#"))

def _font(bold=False, italic=False, size=10, colour="000000"):
    return Font(name="Arial", bold=bold, italic=italic, size=size, color=colour)

def _centre():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

HEADER_FILL  = _fill("#1B4F72")
HEADER_FONT  = _font(bold=True, colour="FFFFFF", size=10)
SUBHEAD_FILL = _fill("#D6EAF8")
SUBHEAD_FONT = _font(bold=True, colour="1B4F72", size=9)
ROW_ODD      = _fill("#F7FBFF")
ROW_EVEN     = _fill("#FFFFFF")
BFA_FILL     = _fill("#FCE4D0")
CTRL_FILL    = _fill("#D5E8F5")
DETECTED_FILL = _fill("#D5F5E3")
PARTIAL_FILL  = _fill("#FEF9E7")
ABSENT_FILL   = _fill("#FDECEA")


def _header_row(ws, row, col_start, col_end, fill=None, font=None, height=22):
    fill = fill or HEADER_FILL
    font = font or HEADER_FONT
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = _centre()
        cell.border = _border()
    ws.row_dimensions[row].height = height


def _data_row(ws, row, col_start, col_end, even=True, fill=None):
    f = fill or (ROW_EVEN if even else ROW_ODD)
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = f
        cell.font = _font(size=9)
        cell.alignment = _left()
        cell.border = _border()


def _col_widths(ws, widths):
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _title(ws, text, row=1, col=1, span=1, size=12):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = _font(bold=True, size=size)
    cell.alignment = _left()
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)


def _note(ws, text, row, col=1, span=1):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = _font(size=8, italic=True, colour="666666")
    cell.alignment = _left()
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)


# ---------------------------------------------------------------------------
# Normalisation pipeline
# ---------------------------------------------------------------------------

def load_and_normalise(path):
    """
    Load raw LC-MS intensity data and apply two-step normalisation.

    Returns
    -------
    df_norm  : DataFrame, gene-level intensities after both normalisation steps
    df_raw   : DataFrame, gene-level raw intensities (pre-normalisation)
    df_step1 : DataFrame, gene-level intensities after Step 1 only
    norm_log : DataFrame, per-fraction normalisation audit table
    """
    raw = pd.read_excel(path)
    raw.columns = raw.columns.str.strip()

    def _gene(desc):
        m = re.search(r"GN=(\S+)", str(desc))
        return m.group(1).upper() if m else ""

    def _name(desc):
        m = re.match(r"^(.+?)\s+OS=", str(desc))
        return m.group(1).strip() if m else str(desc)[:80]

    raw["gene"]      = raw["protein.Description"].apply(_gene)
    raw["prot_name"] = raw["protein.Description"].apply(_name)
    raw["go_slim"]   = raw.apply(
        lambda r: get_go_slim(r["gene"], r["protein.Description"]), axis=1)
    raw["gene_id"]   = raw.apply(
        lambda r: r["gene"] if r["gene"] else r["protein.Accession"], axis=1)

    # Gene-level rollup: sum intensities across paralogs / strain variants
    agg = {c: "sum" for c in QUANT_COLS}
    agg.update({"prot_name": "first", "go_slim": "first",
                "protein.Description": "first", "protein.Accession": "first"})
    df = (raw.groupby("gene_id").agg(agg)
             .reset_index()
             .rename(columns={"gene_id": "gene"}))
    df_raw = df.copy()

    print(f"  {len(raw)} accession-level entries -> {len(df)} gene-level entries")

    # Step 1: loading correction
    norm_log = []
    df_s1 = df.copy()
    for c in QUANT_COLS:
        sf = REFERENCE_UG / PROTEIN_LOADED_UG[c]
        df_s1[c] = df[c].apply(lambda x, sf=sf: x * sf if x > 0 else 0)
        norm_log.append({
            "Fraction":              c,
            "µg loaded":             PROTEIN_LOADED_UG[c],
            "Reference µg":          REFERENCE_UG,
            "Scale factor (Step 1)": round(sf, 4),
            "Median raw":            round(df[df[c] > 0][c].median(), 0),
            "Median after Step 1":   round(df_s1[df_s1[c] > 0][c].median(), 0),
        })

    # Step 2: median normalisation
    gm = np.median([v for c in QUANT_COLS for v in df_s1[c].values if v > 0])
    df_norm = df_s1.copy()
    for i, c in enumerate(QUANT_COLS):
        cm = df_s1[df_s1[c] > 0][c].median()
        df_norm[c] = df_s1[c].apply(
            lambda x, cm=cm: x / cm * gm if x > 0 else 0)
        norm_log[i]["Global median (post Step 1)"] = round(gm, 0)
        norm_log[i]["Median after Step 2"] = round(
            df_norm[df_norm[c] > 0][c].median(), 0)

    return df_norm, df_raw, df_s1, pd.DataFrame(norm_log)


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def write_norm_log(wb, norm_log_df):
    ws = wb.create_sheet("0_Normalisation_Log")
    ws.sheet_view.showGridLines = False
    _title(ws, "Normalisation audit — two-step intensity correction", 1, 1, 8, 13)
    _note(ws,
        "Step 1: raw intensity / µg loaded × 50 µg reference. "
        "Step 2: scale each fraction to the global median across all fractions.",
        2, 1, 8)
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 32

    _header_row(ws, 4, 1, len(norm_log_df.columns))
    for c, h in enumerate(norm_log_df.columns, start=1):
        ws.cell(row=4, column=c, value=h)

    for r_idx, row in norm_log_df.iterrows():
        er = 5 + r_idx
        fill = BFA_FILL if "BfA" in str(row["Fraction"]) else CTRL_FILL
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=er, column=c, value=val)
            cell.fill = fill
            cell.font = _font(size=9)
            cell.alignment = _centre()
            cell.border = _border()
            if isinstance(val, float) and val > 100:
                cell.number_format = "#,##0"

    _col_widths(ws, [22, 12, 14, 20, 16, 20, 26, 20])


def write_venn(wb, df):
    ws = wb.create_sheet("A_Venn")
    ws.sheet_view.showGridLines = False

    sL = set(df.loc[df["lysate"] > 0,      "gene"])
    sS = set(df.loc[df["supernatant"] > 0, "gene"])
    sM = set(df.loc[df["microsomes"] > 0,  "gene"])

    regions = {
        "Lysate only":              sorted(sL - sS - sM),
        "Supernatant only":         sorted(sS - sL - sM),
        "Microsomes only":          sorted(sM - sL - sS),
        "Lysate ∩ Supernatant":     sorted((sL & sS) - sM),
        "Lysate ∩ Microsomes":      sorted((sL & sM) - sS),
        "Supernatant ∩ Microsomes": sorted((sS & sM) - sL),
        "All three fractions":      sorted(sL & sS & sM),
    }

    _title(ws, "Figure 2A — protein detection overlap (control fractions)", 1, 1, 4, 13)
    _note(ws, f"Lysate n={len(sL)}, Supernatant n={len(sS)}, "
              f"Microsomes n={len(sM)}. Total unique: {len(sL | sS | sM)}.",
          2, 1, 4)
    ws.row_dimensions[2].height = 28

    _header_row(ws, 4, 1, 2, fill=SUBHEAD_FILL, font=SUBHEAD_FONT)
    ws.cell(row=4, column=1, value="Venn region")
    ws.cell(row=4, column=2, value="n proteins")
    for i, (region, genes) in enumerate(regions.items()):
        ws.cell(row=5 + i, column=1, value=region)
        ws.cell(row=5 + i, column=2, value=len(genes))
        _data_row(ws, 5 + i, 1, 2, even=(i % 2 == 0))

    start_col = 4
    _header_row(ws, 4, start_col, start_col + len(regions) - 1)
    for c, h in enumerate(regions.keys(), start=start_col):
        ws.cell(row=4, column=c, value=h)

    max_len = max(len(v) for v in regions.values())
    for r in range(max_len):
        for c, genes in enumerate(regions.values(), start=start_col):
            val = genes[r] if r < len(genes) else ""
            cell = ws.cell(row=5 + r, column=c, value=val)
            cell.font = _font(size=9, italic=bool(val))
            cell.alignment = _left()
            cell.border = _border()
            cell.fill = ROW_EVEN if r % 2 == 0 else ROW_ODD

    _col_widths(ws, [28, 12, 4] + [18] * 7)


def write_functional_comp(wb, df):
    ws = wb.create_sheet("B_Functional_Comp")
    ws.sheet_view.showGridLines = False

    cat_order = [
        "Translation/Ribosomal", "Energy metabolism",
        "Chaperones/Protein folding", "Amino acid metabolism",
        "Redox homeostasis", "Proteolysis/Proteasome",
        "ER/Golgi/Secretory", "Nucleic acid metabolism",
        "Uncharacterised", "Other",
    ]
    fracs = [("Lysate", "lysate"), ("Supernatant", "supernatant"),
             ("Microsomes", "microsomes")]

    _title(ws, "Figure 2B — functional composition by fraction (% normalised intensity)",
           1, 1, 5, 13)
    _note(ws,
        "Intensity-weighted %: fraction intensity per GO Slim category / total fraction "
        "intensity. Only proteins with intensity > 0 are included.",
        2, 1, 5)
    ws.row_dimensions[2].height = 28

    _header_row(ws, 4, 1, 5)
    ws.cell(row=4, column=1, value="GO Slim category")
    for c, (label, _) in enumerate(fracs, start=2):
        ws.cell(row=4, column=c, value=f"{label}\n(% intensity)")
    ws.cell(row=4, column=5, value="Colour (hex)")

    frac_totals = {col: df[df[col] > 0][col].sum() for _, col in fracs}

    for i, cat in enumerate(cat_order):
        er = 5 + i
        pcts = []
        for _, col in fracs:
            s = df[(df[col] > 0) & (df["go_slim"] == cat)][col].sum()
            pcts.append(round(s / frac_totals[col] * 100, 2)
                        if frac_totals[col] > 0 else 0)

        ws.cell(er, 1, cat)
        for c, pct in enumerate(pcts, start=2):
            ws.cell(er, c, pct)

        hex_col = CAT_COLOURS.get(cat, "95A5A6")
        ws.cell(er, 5, f"#{hex_col}")
        fill = _fill(hex_col)
        white_cats = ("Translation/Ribosomal", "Amino acid metabolism",
                      "Redox homeostasis")
        for c in range(1, 6):
            cell = ws.cell(er, c)
            cell.fill = fill
            cell.font = _font(size=9, colour="FFFFFF" if cat in white_cats else "000000")
            cell.alignment = _centre() if c > 1 else _left()
            cell.border = _border()
            if c in (2, 3, 4):
                cell.number_format = "0.00"

    tot_row = 5 + len(cat_order)
    ws.cell(tot_row, 1, "TOTAL").font = _font(bold=True, size=9)
    for c in range(2, 5):
        col_letter = get_column_letter(c)
        ws.cell(tot_row, c, f"=SUM({col_letter}5:{col_letter}{tot_row - 1})")
        ws.cell(tot_row, c).number_format = "0.00"
    _data_row(ws, tot_row, 1, 5, fill=SUBHEAD_FILL)
    ws.cell(tot_row, 1).font = SUBHEAD_FONT

    _col_widths(ws, [32, 18, 18, 18, 22])


def write_top15(wb, df, fraction_col, sheet_name, panel_letter, label):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    top = (df[df[fraction_col] > 0]
           .sort_values(fraction_col, ascending=False)
           .head(15)
           .reset_index(drop=True))

    _title(ws, f"Figure 2{panel_letter} — top 15 proteins by abundance: {label}",
           1, 1, 6, 13)
    _note(ws,
        f"Ranked by normalised intensity ({fraction_col}). "
        "Intensities reflect two-step normalisation (loading correction + median).",
        2, 1, 6)
    ws.row_dimensions[2].height = 28

    cols = ["Rank", "Gene", "Protein name", "GO Slim category",
            f"Normalised intensity ({fraction_col})", "Intensity (×10⁶)"]
    _header_row(ws, 4, 1, len(cols))
    for c, h in enumerate(cols, start=1):
        ws.cell(4, c, h)

    for i, row in top.iterrows():
        er = 5 + i
        intensity = row[fraction_col]
        cat = row["go_slim"]
        hex_col = CAT_COLOURS.get(cat, "95A5A6")
        vals = [i + 1, row["gene"], row["prot_name"], cat,
                round(intensity, 0), round(intensity / 1e6, 3)]

        for c, v in enumerate(vals, start=1):
            cell = ws.cell(er, c, v)
            cell.font = _font(size=9, italic=(c == 2), bold=(c == 2))
            cell.alignment = _centre() if c not in (2, 3) else _left()
            cell.border = _border()
            cell.fill = ROW_EVEN if i % 2 == 0 else ROW_ODD

        ws.cell(er, 4).fill = _fill(hex_col)
        white_cats = ("Translation/Ribosomal", "Amino acid metabolism",
                      "Redox homeostasis")
        ws.cell(er, 4).font = _font(
            size=9, colour="FFFFFF" if cat in white_cats else "000000")
        ws.cell(er, 5).number_format = "#,##0"
        ws.cell(er, 6).number_format = "0.000"

    _col_widths(ws, [8, 12, 42, 30, 28, 16])


def write_secretory(wb, df):
    ws = wb.create_sheet("E_Secretory_Pathway")
    ws.sheet_view.showGridLines = False

    _title(ws, "Table 1 — secretory pathway protein detection (control condition)",
           1, 1, 9, 13)
    _note(ws,
        "Detection defined as normalised intensity > 0. 39 proteins spanning "
        "nine functional groups. Supernatant included for reference.",
        2, 1, 9)
    ws.row_dimensions[2].height = 32

    headers = ["Group", "Gene", "Protein / complex role",
               "Lysate\n(detected)", "Supernatant\n(detected)",
               "Microsomes\n(detected)",
               "Lysate intensity", "Supernatant intensity", "Microsomes intensity"]
    _header_row(ws, 4, 1, len(headers))
    for c, h in enumerate(headers, start=1):
        ws.cell(4, c, h)

    prev_group = None
    for i, (gene, display, group) in enumerate(SECRETORY_PROTEINS):
        er = 5 + i
        match = df[df["gene"].str.upper() == gene.upper()]
        if len(match) > 0:
            row  = match.iloc[0]
            lys_i = round(row["lysate"], 0)
            sup_i = round(row["supernatant"], 0)
            mic_i = round(row["microsomes"], 0)
        else:
            lys_i = sup_i = mic_i = 0

        in_lys = lys_i > 0
        in_sup = sup_i > 0
        in_mic = mic_i > 0
        bg = DETECTED_FILL if (in_lys and in_mic) else (
             PARTIAL_FILL if (in_lys or in_mic) else ABSENT_FILL)
        group_val = group if group != prev_group else ""
        gc = GROUP_COLOURS.get(group, "95A5A6")

        vals = [group_val, gene, display,
                "✓" if in_lys else "✗",
                "✓" if in_sup else "✗",
                "✓" if in_mic else "✗",
                lys_i or "", sup_i or "", mic_i or ""]

        for c, v in enumerate(vals, start=1):
            cell = ws.cell(er, c, v)
            cell.border = _border()
            cell.alignment = _centre() if c not in (1, 2, 3) else _left()
            cell.font = _font(size=9, bold=(c == 1 and bool(group_val)),
                              italic=(c == 2))
            if c == 1 and group_val:
                cell.fill = _fill(gc)
                cell.font = _font(bold=True, size=9, colour="FFFFFF")
            elif c in (4, 5, 6):
                cell.fill = DETECTED_FILL if v == "✓" else ABSENT_FILL
                cell.font = _font(bold=True, size=10,
                                  colour="1E8449" if v == "✓" else "C0392B")
            else:
                cell.fill = bg
                if c in (7, 8, 9) and v:
                    cell.number_format = "#,##0"

        if group != prev_group and prev_group is not None:
            thick = Side(style="medium", color="888888")
            for c in range(1, 10):
                ex = ws.cell(er, c).border
                ws.cell(er, c).border = Border(
                    left=ex.left, right=ex.right, bottom=ex.bottom, top=thick)

        prev_group = group

    leg_row = 5 + len(SECRETORY_PROTEINS) + 1
    ws.cell(leg_row, 1, "Legend:").font = _font(bold=True, size=9)
    for j, (fill, text) in enumerate([
        (DETECTED_FILL, "✓ detected in both lysate and microsomes"),
        (PARTIAL_FILL,  "partial — detected in one fraction only"),
        (ABSENT_FILL,   "✗ not detected in either fraction"),
    ]):
        ws.cell(leg_row + 1 + j, 1, text).fill = fill
        ws.cell(leg_row + 1 + j, 1).font = _font(size=9)
        ws.cell(leg_row + 1 + j, 1).border = _border()

    _col_widths(ws, [22, 10, 36, 14, 16, 14, 20, 22, 20])


def write_all_proteins(wb, df):
    ws = wb.create_sheet("All_Proteins")
    ws.sheet_view.showGridLines = False

    _title(ws, "Full normalised protein table — all 618 gene-level entries",
           1, 1, 12, 13)
    _note(ws,
        "Two-step normalised intensities. Zeros indicate the protein was not detected "
        "in that fraction. BfA-exclusive = detected in at least one BfA fraction but "
        "absent from all control fractions.",
        2, 1, 12)
    ws.row_dimensions[2].height = 28

    cols = ["Gene", "Protein name", "GO Slim category",
            "BfA Lysate", "BfA Supernatant", "BfA Microsomes",
            "Ctrl Lysate", "Ctrl Supernatant", "Ctrl Microsomes",
            "Detected in\nany BfA", "Detected in\nany Ctrl", "BfA-exclusive"]
    _header_row(ws, 4, 1, len(cols))
    for c, h in enumerate(cols, start=1):
        ws.cell(4, c, h)

    export = df[["gene", "prot_name", "go_slim"] + QUANT_COLS].copy()
    export.columns = ["Gene", "Protein name", "GO Slim category",
                      "BfA Lysate", "BfA Supernatant", "BfA Microsomes",
                      "Ctrl Lysate", "Ctrl Supernatant", "Ctrl Microsomes"]
    export = export.sort_values("GO Slim category").reset_index(drop=True)

    bfa_intensity_cols  = ["BfA Lysate", "BfA Supernatant", "BfA Microsomes"]
    ctrl_intensity_cols = ["Ctrl Lysate", "Ctrl Supernatant", "Ctrl Microsomes"]
    white_cats = ("Translation/Ribosomal", "Amino acid metabolism", "Redox homeostasis")

    for i, row in export.iterrows():
        er = 5 + i
        even = i % 2 == 0
        cat = row["GO Slim category"]
        hex_col = CAT_COLOURS.get(cat, "95A5A6")
        bfa_any  = any(row[c] > 0 for c in bfa_intensity_cols)
        ctrl_any = any(row[c] > 0 for c in ctrl_intensity_cols)
        bfa_excl = bfa_any and not ctrl_any

        vals = list(row) + [
            "Yes" if bfa_any  else "No",
            "Yes" if ctrl_any else "No",
            "Yes" if bfa_excl else "No",
        ]

        for c, v in enumerate(vals, start=1):
            cell = ws.cell(er, c, v if v != 0.0 else "")
            cell.border = _border()
            cell.font = _font(size=8, italic=(c == 1))
            cell.alignment = _left() if c <= 3 else _centre()
            if c == 3:
                cell.fill = _fill(hex_col)
                cell.font = _font(size=8,
                    colour="FFFFFF" if cat in white_cats else "000000")
            elif c in range(4, 10):
                cell.fill = (DETECTED_FILL if (v and v > 0)
                             else (ROW_EVEN if even else ROW_ODD))
                if v and v > 0:
                    cell.number_format = "#,##0"
            elif c == 12 and v == "Yes":
                cell.fill = BFA_FILL
                cell.font = _font(size=8, bold=True)
            else:
                cell.fill = ROW_EVEN if even else ROW_ODD

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:L4"
    _col_widths(ws, [14, 40, 28, 14, 18, 16, 14, 18, 16, 14, 14, 14])


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    print("Loading data and applying normalisation ...")
    df_norm, df_raw, df_s1, norm_log = load_and_normalise(DATA_PATH)

    print("\nBuilding output workbook ...")
    wb = Workbook()
    wb.remove(wb.active)

    write_norm_log(wb, norm_log);       print("  0_Normalisation_Log")
    write_venn(wb, df_norm);            print("  A_Venn")
    write_functional_comp(wb, df_norm); print("  B_Functional_Comp")
    write_top15(wb, df_norm, "lysate",      "C_Top15_Lysate",      "C", "control lysate")
    print("  C_Top15_Lysate")
    write_top15(wb, df_norm, "microsomes",  "D_Top15_Microsomes",  "D", "control microsomes")
    print("  D_Top15_Microsomes")
    write_secretory(wb, df_norm);       print("  E_Secretory_Pathway")
    write_all_proteins(wb, df_norm);    print("  All_Proteins")

    wb.save(OUT_PATH)
    print(f"\nSaved: {OUT_PATH}")


if __name__ == "__main__":
    main()
